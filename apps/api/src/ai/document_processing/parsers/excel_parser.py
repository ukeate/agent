"""Excel文档解析器"""

import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
import json

import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd

from .base_parser import BaseParser, ParsedDocument, ParsedElement

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    """Excel文档解析器
    
    支持.xlsx和.xls格式的表格数据、公式、图表提取
    """
    
    SUPPORTED_EXTENSIONS = [".xlsx", ".xls", ".csv"]
    
    def __init__(
        self, 
        extract_formulas: bool = True,
        extract_charts: bool = True,
        max_rows: Optional[int] = None
    ):
        """初始化Excel解析器
        
        Args:
            extract_formulas: 是否提取公式
            extract_charts: 是否提取图表信息
            max_rows: 最大读取行数限制
        """
        super().__init__()
        self.extract_formulas = extract_formulas
        self.extract_charts = extract_charts
        self.max_rows = max_rows
    
    async def parse(self, file_path: Path) -> ParsedDocument:
        """解析Excel文档
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            解析后的文档
        """
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        doc_id = self.generate_doc_id(file_path)
        metadata = self.extract_metadata(file_path)
        elements = []
        
        # 根据文件类型选择解析方法
        if file_path.suffix.lower() == ".csv":
            elements = await self._parse_csv(file_path)
        else:
            elements = await self._parse_excel(file_path)
        
        return ParsedDocument(
            doc_id=doc_id,
            file_path=str(file_path),
            file_type="spreadsheet",
            elements=elements,
            metadata=metadata
        )
    
    async def _parse_excel(self, file_path: Path) -> List[ParsedElement]:
        """解析Excel文件(.xlsx/.xls)
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            解析的元素列表
        """
        elements = []
        
        try:
            # 加载工作簿
            workbook = openpyxl.load_workbook(
                str(file_path), 
                data_only=not self.extract_formulas
            )
            
            # 遍历所有工作表
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # 提取工作表数据
                sheet_element = await self._parse_worksheet(sheet, sheet_name)
                elements.append(sheet_element)
                
                # 提取公式
                if self.extract_formulas:
                    formulas = await self._extract_formulas(sheet, sheet_name)
                    if formulas:
                        elements.append(formulas)
                
                # 提取图表
                if self.extract_charts and hasattr(sheet, '_charts'):
                    charts = await self._extract_charts(sheet, sheet_name)
                    elements.extend(charts)
            
            workbook.close()
            
        except Exception as e:
            logger.error(f"Error parsing Excel file: {e}")
            # 降级使用pandas
            elements = await self._parse_with_pandas(file_path)
        
        return elements
    
    async def _parse_worksheet(
        self, 
        sheet, 
        sheet_name: str
    ) -> ParsedElement:
        """解析工作表
        
        Args:
            sheet: 工作表对象
            sheet_name: 工作表名称
            
        Returns:
            解析的元素
        """
        # 获取数据范围
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # 应用行数限制
        if self.max_rows and max_row > self.max_rows:
            max_row = self.max_rows
        
        # 提取数据
        data = []
        headers = []
        
        for row_idx in range(1, max_row + 1):
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = cell.value
                
                # 处理不同类型的单元格值
                if value is None:
                    value = ""
                elif isinstance(value, (int, float)):
                    value = str(value)
                elif hasattr(value, 'isoformat'):  # datetime
                    value = value.isoformat()
                else:
                    value = str(value)
                
                row_data.append(value)
            
            if row_idx == 1:
                headers = row_data
            else:
                data.append(row_data)
        
        # 转换为表格格式
        content_lines = []
        
        # 添加表头
        if headers:
            content_lines.append(" | ".join(headers))
            content_lines.append("-" * (len(" | ".join(headers))))
        
        # 添加数据行
        for row in data[:100]:  # 限制显示前100行
            content_lines.append(" | ".join(row))
        
        if len(data) > 100:
            content_lines.append(f"... and {len(data) - 100} more rows")
        
        content = "\n".join(content_lines)
        
        # 计算统计信息
        stats = self._calculate_sheet_stats(sheet)
        
        return ParsedElement(
            content=content,
            element_type="table",
            metadata={
                "sheet_name": sheet_name,
                "headers": headers,
                "num_rows": len(data) + 1,  # 包括标题行
                "num_columns": max_col,
                "stats": stats,
                "has_formulas": any(
                    sheet.cell(r, c).data_type == 'f' 
                    for r in range(1, min(10, max_row + 1)) 
                    for c in range(1, max_col + 1)
                )
            }
        )
    
    async def _extract_formulas(self, sheet, sheet_name: str) -> Optional[ParsedElement]:
        """提取工作表中的公式
        
        Args:
            sheet: 工作表对象
            sheet_name: 工作表名称
            
        Returns:
            公式元素或None
        """
        formulas = []
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # 公式类型
                    formulas.append({
                        "cell": f"{get_column_letter(cell.column)}{cell.row}",
                        "formula": cell.value,
                        "result": sheet.cell(
                            row=cell.row, 
                            column=cell.column
                        ).value
                    })
        
        if not formulas:
            return None
        
        # 格式化公式内容
        content_lines = ["Formulas in sheet:"]
        for f in formulas[:20]:  # 限制显示前20个公式
            content_lines.append(
                f"{f['cell']}: {f['formula']} = {f['result']}"
            )
        
        if len(formulas) > 20:
            content_lines.append(f"... and {len(formulas) - 20} more formulas")
        
        return ParsedElement(
            content="\n".join(content_lines),
            element_type="formula",
            metadata={
                "sheet_name": sheet_name,
                "num_formulas": len(formulas),
                "formulas": formulas[:20]  # 存储前20个公式的详细信息
            }
        )
    
    async def _extract_charts(self, sheet, sheet_name: str) -> List[ParsedElement]:
        """提取工作表中的图表
        
        Args:
            sheet: 工作表对象
            sheet_name: 工作表名称
            
        Returns:
            图表元素列表
        """
        elements = []
        
        try:
            for idx, chart in enumerate(sheet._charts):
                chart_info = {
                    "type": chart.__class__.__name__,
                    "title": chart.title if hasattr(chart, 'title') else f"Chart {idx + 1}",
                }
                
                # 提取图表数据范围
                if hasattr(chart, 'series'):
                    chart_info["series_count"] = len(chart.series)
                
                elements.append(ParsedElement(
                    content=f"Chart: {chart_info['title']} (Type: {chart_info['type']})",
                    element_type="chart",
                    metadata={
                        "sheet_name": sheet_name,
                        "chart_index": idx,
                        **chart_info
                    }
                ))
        
        except Exception as e:
            logger.warning(f"Error extracting charts: {e}")
        
        return elements
    
    async def _parse_csv(self, file_path: Path) -> List[ParsedElement]:
        """解析CSV文件
        
        Args:
            file_path: CSV文件路径
            
        Returns:
            解析的元素列表
        """
        elements = []
        
        try:
            # 使用pandas读取CSV
            df = pd.read_csv(file_path, nrows=self.max_rows)
            
            # 转换为表格格式
            content_lines = []
            
            # 添加表头
            headers = df.columns.tolist()
            content_lines.append(" | ".join(map(str, headers)))
            content_lines.append("-" * (len(" | ".join(map(str, headers)))))
            
            # 添加数据行
            for _, row in df.head(100).iterrows():
                content_lines.append(" | ".join(map(str, row.values)))
            
            if len(df) > 100:
                content_lines.append(f"... and {len(df) - 100} more rows")
            
            content = "\n".join(content_lines)
            
            # 计算统计信息
            stats = {
                "numeric_columns": df.select_dtypes(include=['number']).columns.tolist(),
                "null_counts": df.isnull().sum().to_dict(),
                "shape": df.shape,
            }
            
            elements.append(ParsedElement(
                content=content,
                element_type="table",
                metadata={
                    "headers": headers,
                    "num_rows": len(df),
                    "num_columns": len(df.columns),
                    "stats": stats,
                }
            ))
            
        except Exception as e:
            logger.error(f"Error parsing CSV file: {e}")
            raise
        
        return elements
    
    async def _parse_with_pandas(self, file_path: Path) -> List[ParsedElement]:
        """使用pandas解析Excel文件（降级方案）
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            解析的元素列表
        """
        elements = []
        
        try:
            # 读取所有工作表
            excel_file = pd.ExcelFile(file_path)
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(
                    excel_file, 
                    sheet_name=sheet_name,
                    nrows=self.max_rows
                )
                
                # 转换为表格格式
                content_lines = []
                
                # 添加表头
                headers = df.columns.tolist()
                content_lines.append(" | ".join(map(str, headers)))
                content_lines.append("-" * (len(" | ".join(map(str, headers)))))
                
                # 添加数据行
                for _, row in df.head(100).iterrows():
                    content_lines.append(" | ".join(map(str, row.values)))
                
                if len(df) > 100:
                    content_lines.append(f"... and {len(df) - 100} more rows")
                
                content = "\n".join(content_lines)
                
                elements.append(ParsedElement(
                    content=content,
                    element_type="table",
                    metadata={
                        "sheet_name": sheet_name,
                        "headers": headers,
                        "num_rows": len(df),
                        "num_columns": len(df.columns),
                    }
                ))
            
        except Exception as e:
            logger.error(f"Error parsing with pandas: {e}")
            raise
        
        return elements
    
    def _calculate_sheet_stats(self, sheet) -> Dict[str, Any]:
        """计算工作表统计信息
        
        Args:
            sheet: 工作表对象
            
        Returns:
            统计信息字典
        """
        stats = {
            "merged_cells": len(sheet.merged_cells.ranges),
            "has_filters": bool(sheet.auto_filter),
        }
        
        # 统计数据类型
        type_counts = {"numeric": 0, "text": 0, "date": 0, "formula": 0, "empty": 0}
        
        sample_size = min(100, sheet.max_row)
        for row in sheet.iter_rows(max_row=sample_size):
            for cell in row:
                if cell.value is None:
                    type_counts["empty"] += 1
                elif cell.data_type == 'f':
                    type_counts["formula"] += 1
                elif cell.data_type == 'n':
                    type_counts["numeric"] += 1
                elif cell.data_type == 'd':
                    type_counts["date"] += 1
                else:
                    type_counts["text"] += 1
        
        stats["type_distribution"] = type_counts
        
        return stats